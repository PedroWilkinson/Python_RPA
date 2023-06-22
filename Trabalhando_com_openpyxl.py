from openpyxl import load_workbook
import os

from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell

caminho_nome_arquivo = "C:\\Users\\Pedro W\\Desktop\\PYTHON_GERAL\\PythonRPA\\Openpyxl\\Vendedores.xlsx"
planilha_aberta = load_workbook(filename=caminho_nome_arquivo)

#Seleciona a sheet de Vendas
sheet_selecionada = planilha_aberta['Vendas']

somarAmandaMartins = 0
somarElianeMoreira = 0
somarLeonardoAlmeida = 0
somarNicolasPereira = 0

for linha in range(2, len(sheet_selecionada['A']) + 1):

    if sheet_selecionada['A%s' % linha].value == "Amanda Martins":
        somarAmandaMartins = somarAmandaMartins + sheet_selecionada['C%s' % linha].value
    
    elif sheet_selecionada['A%s' % linha].value == "Eliane Moreira":
        somarElianeMoreira = somarElianeMoreira + sheet_selecionada['C%s' % linha].value
    
    elif sheet_selecionada['A%s' % linha].value == "Nicolas Pereira":
        somarNicolasPereira = somarNicolasPereira + sheet_selecionada['C%s' % linha].value

    elif sheet_selecionada['A%s' % linha].value == "Leonardo Almeida":
        somarLeonardoAlmeida = somarLeonardoAlmeida + sheet_selecionada['C%s' % linha].value

sheet_resumo = planilha_aberta.create_sheet(title="Resumo")

sheet_resumo['A1'] = "Vendedores"
sheet_resumo['B1'] = "Vendas"

sheet_resumo['A2'] = "Amanda Martins"
sheet_resumo['B2'] = somarAmandaMartins

sheet_resumo['A3'] = "Eliane Moreira"
sheet_resumo['B3'] = somarElianeMoreira

sheet_resumo['A4'] = "Nicolas Pereira"
sheet_resumo['B4'] = somarNicolasPereira

sheet_resumo['A5'] = "Leonardo Almeida"
sheet_resumo['B5'] = somarLeonardoAlmeida

sheet_resumo['A6'] = "Leonardo Almeida"
sheet_resumo['B6'] = somarLeonardoAlmeida


    
'''


sheet_selecionada['A6'] = "=SOMA(A2:A5)" 
sheet_selecionada['B6'] = "=SOMA(B2:B5)"
sheet_selecionada['D2'] = "=A2+B2"
sheet_selecionada['D3'] = "=A3-B3"
sheet_selecionada['D4'] = "=A4*B4"
sheet_selecionada['D5'] = "=A5/B5"


sheet_selecionada['B12'] = "=MID(A12,1,3)"
sheet_selecionada['C12'] = "=MID(A12,5,3)"
sheet_selecionada['D12'] = "=MID(A12,9,3)"
sheet_selecionada['E12'] = "=MID(A12,13,2)"




#Popula as informacoes que vao para a planilha
dadosTabela = [
    ['Nome', 'Idade'],
    ['Berenice', 28],
    ['Caio', 32],
    ['Nicole', 34]
]

#O Append pega toda a lista e passa parar a sheet
for linhaPlanilha in dadosTabela:
    sheet_selecionada.append(linhaPlanilha)


corTitulo = PatternFill(start_color= '00FFFF00',
                       end_color= '00FFFF00',
                       fill_type='solid')

corCelulas = PatternFill(start_color= '00CCFFCC',
                       end_color= '00CCFFCC',
                       fill_type='solid')

sheet_selecionada["A1"].fill = corTitulo
sheet_selecionada["B1"].fill = corTitulo

for linha in range(2, len(sheet_selecionada['A']) + 1):

    celulaColunaA = "A" + str(linha)
    celulaColunaB = "B" + str(linha)

    sheet_selecionada[celulaColunaA].fill = corCelulas
    sheet_selecionada[celulaColunaB].fill = corCelulas


#Deleta as Linhas
sheet_selecionada.delete_rows(3)
sheet_selecionada.delete_rows(3)
sheet_selecionada.delete_rows(5)

#Deleta a Coluna
sheet_selecionada.delete_cols(2)
'''

#Salva a planilha com as alteracoes
planilha_aberta.save(filename=caminho_nome_arquivo)

#Abre a planilha
os.startfile(caminho_nome_arquivo)
